import streamlit as st
import pandas as pd
import numpy as np
import os
import json
import requests
import time
from io import BytesIO
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Sayfa yapılandırması
st.set_page_config(
    page_title="E-Ticaret Fiyatlandırma & Satış Otomasyonu",
    page_icon="🔐",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Sabit Dosya Yolları
DB_FILE = 'urun_maliyet_veritabani.csv'
API_FILE = 'api_ayarlari.json'
AUTH_FILE = 'auth_config.json'

# --- KULLANICI GİRİŞ (LOGIN) & CAPTCHA SİSTEMİ ---
def load_auth():
    default_auth = {"users": {"aliozer73": "Ayten136"}}
    if os.path.exists(AUTH_FILE):
        try:
            with open(AUTH_FILE, 'r', encoding='utf-8') as f:
                saved = json.load(f)
                if "username" in saved and "password" in saved and "users" not in saved:
                    default_auth["users"][saved["username"]] = saved["password"]
                elif "users" in saved and isinstance(saved["users"], dict):
                    default_auth["users"].update(saved["users"])
        except Exception:
            pass
    else:
        with open(AUTH_FILE, 'w', encoding='utf-8') as f:
            json.dump(default_auth, f, ensure_ascii=False, indent=4)
    return default_auth

def save_auth(auth_data):
    with open(AUTH_FILE, 'w', encoding='utf-8') as f:
        json.dump(auth_data, f, ensure_ascii=False, indent=4)

if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
if "current_user" not in st.session_state:
    st.session_state["current_user"] = ""

if "captcha_num1" not in st.session_state or "captcha_num2" not in st.session_state:
    st.session_state["captcha_num1"] = np.random.randint(1, 10)
    st.session_state["captcha_num2"] = np.random.randint(1, 10)

def reset_captcha():
    st.session_state["captcha_num1"] = np.random.randint(1, 10)
    st.session_state["captcha_num2"] = np.random.randint(1, 10)

auth_data = load_auth()

# Özel CSS
st.markdown("""
    <style>
    .main-title { font-size: 30px; font-weight: bold; color: #E74C3C; margin-bottom: 5px; text-shadow: 1px 1px 2px rgba(0,0,0,0.1); }
    .sub-title { font-size: 14px; color: #2E4053; margin-bottom: 25px; }
    .metric-box { background-color: #fcfdfc; padding: 15px; border-radius: 10px; border-left: 6px solid #2ECC71; box-shadow: 0 4px 6px rgba(0,0,0,0.05); }
    .hb-title { font-size: 30px; font-weight: bold; color: #FF6700; margin-bottom: 5px; text-shadow: 1px 1px 2px rgba(0,0,0,0.1); }
    .hb-metric { background-color: #fcfdfc; padding: 15px; border-radius: 10px; border-left: 6px solid #FF6700; box-shadow: 0 4px 6px rgba(0,0,0,0.05); }
    .sales-title { font-size: 30px; font-weight: bold; color: #2980B9; margin-bottom: 5px; text-shadow: 1px 1px 2px rgba(0,0,0,0.1); }
    .sales-metric { background-color: #f8fbfe; padding: 15px; border-radius: 10px; border-left: 6px solid #2980B9; box-shadow: 0 4px 6px rgba(0,0,0,0.05); }
    .calc-title { font-size: 30px; font-weight: bold; color: #8E44AD; margin-bottom: 5px; text-shadow: 1px 1px 2px rgba(0,0,0,0.1); }
    .calc-metric { background-color: #fdfafc; padding: 15px; border-radius: 10px; border-left: 6px solid #8E44AD; box-shadow: 0 4px 6px rgba(0,0,0,0.05); }
    .highlight-card { background: linear-gradient(135deg, #f6f9fc 0%, #edf2f7 100%); padding: 15px; border-radius: 10px; border: 1px solid #cbd5e0; text-align: center; }
    .login-box { max-width: 400px; margin: 100px auto; padding: 30px; background: #ffffff; border-radius: 12px; box-shadow: 0 8px 24px rgba(0,0,0,0.1); border: 1px solid #e2e8f0; }
    .stButton>button { background-color: #2ECC71; color: white; border-radius: 8px; border: none; font-weight: bold; transition: all 0.3s; }
    .stButton>button:hover { background-color: #27AE60; transform: translateY(-2px); box-shadow: 0 4px 8px rgba(46, 204, 113, 0.4); }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] { background-color: #f1f3f5; border-radius: 4px 4px 0 0; padding: 10px 20px; font-weight: 600; color: #495057; }
    .stTabs [aria-selected="true"] { background-color: #2ECC71 !important; color: white !important; }
    </style>
""", unsafe_allow_html=True)

if not st.session_state["logged_in"]:
    st.markdown("<h1 style='text-align: center; color: #2C3E50; margin-top: 50px;'>🔐 E-Ticaret Otomasyon Girişi</h1>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        with st.form("login_form"):
            st.markdown("### Lütfen Giriş Yapın")
            kadi = st.text_input("Kullanıcı Adı")
            sifre = st.text_input("Şifre", type="password")
            st.markdown("---")
            st.markdown("#### 🛡 Güvenlik Doğrulaması")
            num1 = st.session_state["captcha_num1"]
            num2 = st.session_state["captcha_num2"]
            captcha_ans = st.text_input(f"🤖 Gerçek insan olduğunuzu doğrulayın: {num1} + {num2} kaçtır?", placeholder="Sonucu buraya yazın")
            submit = st.form_submit_button("🚀 Sisteme Giriş Yap", use_container_width=True)
            if submit:
                try:
                    user_ans = int(captcha_ans.strip())
                except ValueError:
                    user_ans = -1
                if user_ans != (num1 + num2):
                    st.error("❌ Güvenlik sorusunu (doğrulama kodunu) yanlış cevapladınız. Lütfen tekrar deneyin!")
                    reset_captcha()
                elif kadi.strip() in auth_data["users"] and auth_data["users"][kadi.strip()] == sifre.strip():
                    st.session_state["logged_in"] = True
                    st.session_state["current_user"] = kadi.strip()
                    st.success("✅ Giriş Başarılı! Yönlendiriliyorsunuz...")
                    st.rerun()
                else:
                    st.error("❌ Kullanıcı adı veya şifre hatalı!")
                    reset_captcha()
    st.stop()

# --- YARDIMCI FONKSİYONLAR ---
def turkce_tarih_format(dt_obj):
    if pd.isna(dt_obj):
        return ""
    aylar = {1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan", 5: "Mayıs", 6: "Haziran", 7: "Temmuz", 8: "Ağustos", 9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"}
    return f"{dt_obj.day} {aylar.get(dt_obj.month, '')} {dt_obj.year} {dt_obj.strftime('%H:%M')}"

def tablayi_1den_baslat(df):
    df_copy = df.copy()
    df_copy.index = np.arange(1, len(df_copy) + 1)
    df_copy.index.name = "Sıra"
    return df_copy

def temizle_ve_sayiya_donustur(val):
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip()
    if not val_str:
        return 0.0
    if '.' in val_str and ',' in val_str:
        if val_str.rfind('.') > val_str.rfind(','):
            val_str = val_str.replace(',', '')
        else:
            val_str = val_str.replace('.', '').replace(',', '.')
    elif ',' in val_str:
        val_str = val_str.replace(',', '.')
    try:
        return float(val_str)
    except ValueError:
        return 0.0

def load_db():
    if os.path.exists(DB_FILE):
        try:
            return pd.read_csv(DB_FILE, dtype={'Barkod': str})
        except Exception:
            return pd.DataFrame(columns=['Barkod', 'Ürün Adı', 'Maliyet (TL)', 'Kargo (TL)', 'Komisyon (%)'])
    return pd.DataFrame(columns=['Barkod', 'Ürün Adı', 'Maliyet (TL)', 'Kargo (TL)', 'Komisyon (%)'])

def find_default_col(options, keywords, exclude_keywords=None):
    if exclude_keywords is None:
        exclude_keywords = []
    for opt in options:
        opt_lower = str(opt).lower()
        if any(kw in opt_lower for kw in keywords) and not any(ex_kw in opt_lower for ex_kw in exclude_keywords):
            return options.index(opt)
    return 0

def load_api_settings():
    default_settings = {"ty_seller_id": "", "ty_api_key": "", "ty_api_secret": ""}
    if os.path.exists(API_FILE):
        try:
            with open(API_FILE, 'r', encoding='utf-8') as f:
                saved = json.load(f)
                default_settings.update(saved)
        except Exception:
            pass
    return default_settings

def save_api_settings(settings):
    with open(API_FILE, 'w', encoding='utf-8') as f:
        json.dump(settings, f, ensure_ascii=False, indent=4)

# --- TRENDYOL API İSTEK FONKSİYONLARI - DÜZELTİLDİ ---
def ty_api_request(url, method="GET", payload=None):
    api = load_api_settings()
    if not api["ty_seller_id"] or not api["ty_api_key"] or not api["ty_api_secret"]:
        st.error("❌ Trendyol API bilgileri bulunamadı! Lütfen sol menüdeki '⚙ Ayarlar & API' bölümünden bilgilerinizi bir kere kaydedin.")
        return None
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    auth = (api["ty_api_key"], api["ty_api_secret"])
    try:
        if method == "GET":
            response = requests.get(url, headers=headers, auth=auth, timeout=30)
        elif method == "POST":
            response = requests.post(url, headers=headers, auth=auth, json=payload, timeout=30)
        elif method == "PUT":
            response = requests.put(url, headers=headers, auth=auth, json=payload, timeout=30)
        
        if response.status_code in [200, 201]:
            return response.json()
        elif response.status_code == 429:
            st.warning("⏳ Trendyol rate limit (429), 2 saniye bekleniyor...")
            time.sleep(2)
            return None
        elif response.status_code == 403:
            st.error("❌ Trendyol (403 Forbidden): API güvenlik duvarına takıldınız. Lütfen bu kodu kendi bilgisayarınızda (lokal) çalıştırın.")
            return None
        else:
            st.error(f"❌ Trendyol API Hatası ({response.status_code}): {response.text[:500]}")
            return None
    except Exception as e:
        st.error(f"❌ Bağlantı Hatası: {str(e)}")
        return None

def fetch_ty_orders(start_dt, end_dt):
    """DÜZELTİLDİ - Canlı satışları eksiksiz çekmek için pagination ve tarih mantığı tamamen yenilendi"""
    api = load_api_settings()
    all_orders = []
    current_start = start_dt
    
    progress_bar = st.progress(0, text="Siparişler çekiliyor...")
    total_seconds = (end_dt - start_dt).total_seconds()
    if total_seconds <= 0:
        total_seconds = 1

    while current_start < end_dt:
        current_end = min(current_start + timedelta(days=14), end_dt)
        start_ts = int(current_start.timestamp() * 1000)
        end_ts = int(current_end.timestamp() * 1000)
        
        page = 0
        while True:
            url = f"https://api.trendyol.com/sapigw/suppliers/{api['ty_seller_id']}/orders?startDate={start_ts}&endDate={end_ts}&size=200&page={page}"
            data = ty_api_request(url)
            
            if not data:
                time.sleep(1)
                data = ty_api_request(url)
                if not data:
                    break

            content = data.get("content", [])
            if not content:
                break
            
            all_orders.extend(content)
            
            # Düzgün bitiş kontrolü - 2 koşuldan biri olursa son sayfa
            if len(content) < 200:
                break
            total_pages = data.get("totalPages")
            if total_pages is not None and page >= total_pages - 1:
                break
            
            page += 1
            time.sleep(0.35)  # Rate limit koruması
        
        # Progress güncelle
        elapsed = (current_end - start_dt).total_seconds()
        progress_bar.progress(min(elapsed / total_seconds, 1.0), text=f"{len(all_orders)} sipariş çekildi...")
        current_start = current_end
        
    progress_bar.empty()
    return all_orders

# --- YAN MENÜ ---
st.sidebar.markdown("<h2 style='text-align: center; color: #E74C3C;'>🏷 Avantajlı Ürün</h2>", unsafe_allow_html=True)
st.sidebar.markdown("<p style='text-align: center; color: #2ECC71; font-weight: bold;'>Fiyatlandırma & Satış Sistemi</p>", unsafe_allow_html=True)
if st.session_state["current_user"]:
    st.sidebar.markdown(f"<p style='text-align: center; color: #555; font-size: 13px;'>👤 Aktif Kullanıcı: <b>{st.session_state['current_user']}</b></p>", unsafe_allow_html=True)
st.sidebar.markdown("---")
menu = st.sidebar.radio("Sayfa Seçimi:", [
    "📦 Maliyet Yönetimi", 
    "🧮 İdeal Fiyat Hesaplama",
    "🚀 Trendyol Yıldızlı Fiyat", 
    "💜 Hepsiburada Avantajlı Teklif",
    "📊 Trendyol Satış Analizi (API)",
    "⚙ Ayarlar & API"
])
st.sidebar.markdown("---")
if st.sidebar.button("🔒 Çıkış Yap", use_container_width=True):
    st.session_state["logged_in"] = False
    st.session_state["current_user"] = ""
    reset_captcha()
    if "ty_satis_raporu" in st.session_state:
        del st.session_state["ty_satis_raporu"]
    st.rerun()

# ==========================================
# SAYFA 1: MALİYET VERİTABANI YÖNETİMİ
# ==========================================
if menu == "📦 Maliyet Yönetimi":
    st.markdown('<div class="main-title">📦 Ortak Veritabanı Yönetimi</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">Trendyol ve Hepsiburada hesaplamaları için ürün maliyetlerinizi, kargo ve komisyon oranlarınızı buradan düzenleyin.</div>', unsafe_allow_html=True)
    
    mevcut_db = load_db()
    
    if mevcut_db.empty:
        ornek_veri = pd.DataFrame({
            'Barkod': ['ORNEK_BARKOD_1'],
            'Ürün Adı': ['Örnek Ürün'],
            'Maliyet (TL)': [100.0],
            'Kargo (TL)': [45.0],
            'Komisyon (%)': [15.0]
        })
        mevcut_db = pd.concat([mevcut_db, ornek_veri], ignore_index=True)
        st.info("Sistemde ürün yok. Örnek satırın üzerine tıklayarak kendi ürünlerinizi girmeye başlayabilirsiniz.")
        
    edited_df = st.data_editor(
        tablayi_1den_baslat(mevcut_db),
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Barkod": st.column_config.TextColumn("Barkod / SKU", required=True),
            "Ürün Adı": st.column_config.TextColumn("Ürün Adı"),
            "Maliyet (TL)": st.column_config.NumberColumn("Maliyet (TL)", min_value=0.0, format="%.2f"),
            "Kargo (TL)": st.column_config.NumberColumn("Kargo (TL)", min_value=0.0, format="%.2f"),
            "Komisyon (%)": st.column_config.NumberColumn("Komisyon (%)", min_value=0.0, max_value=100.0, format="%.2f"),
        },
        height=500
    )
    
    if st.button("💾 Değişiklikleri Sisteme Kaydet", use_container_width=True):
        df_save = edited_df.reset_index(drop=True)
        df_save['Barkod'] = df_save['Barkod'].astype(str).str.strip()
        df_save = df_save[df_save['Barkod'] != '']
        df_save = df_save[df_save['Barkod'] != 'nan']
        df_save.to_csv(DB_FILE, index=False)
        st.success("✅ Veritabanı başarıyla güncellendi!")

# ==========================================
# SAYFA 2: İDEAL SATIŞ FİYATI HESAPLAMA
# ==========================================
elif menu == "🧮 İdeal Fiyat Hesaplama":
    st.markdown('<div class="calc-title">🧮 İdeal Satış Fiyatı & Kâr Marjı Hesaplayıcı</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">Ürün maliyeti, kargo ve pazaryeri komisyonunu hesaba katarak hedeflediğiniz kâr marjına veya net kâr tutarına ulaşmak için satmanız gereken ideal fiyatı anında hesaplayın.</div>', unsafe_allow_html=True)
    
    tab_tekli, tab_toplu = st.tabs(["⚡ Hızlı Tekli Ürün Hesaplayıcı", "📦 Veritabanındaki Tüm Ürünler İçin Toplu Fiyatlama"])
    
    with tab_tekli:
        st.markdown("#### 🎯 Hızlı Fiyat ve Kâr Simülasyonu")
        col_inp1, col_inp2, col_inp3 = st.columns(3)
        with col_inp1:
            in_maliyet = st.number_input("📦 Ürün Maliyeti (TL)", min_value=0.0, value=150.0, step=5.0, format="%.2f")
            in_kargo = st.number_input("🚚 Kargo Gideri (TL)", min_value=0.0, value=45.0, step=5.0, format="%.2f")
        with col_inp2:
            in_komisyon = st.number_input("🤝 Pazaryeri Komisyonu (%)", min_value=0.0, max_value=95.0, value=18.0, step=1.0, format="%.2f")
            hesap_yontemi = st.radio("Hedefleme Yöntemi:", ["📊 Yüzdelik Kâr Marjı (%)", "💵 Net Kâr Tutarı (TL)"], horizontal=True)
        with col_inp3:
            if "Yüzdelik" in hesap_yontemi:
                in_hedef_val = st.number_input("🎯 Hedef Kâr Marjı (%)", min_value=1.0, max_value=80.0, value=25.0, step=1.0, format="%.2f")
            else:
                in_hedef_val = st.number_input("🎯 Hedef Net Kâr Tutarı (TL)", min_value=1.0, value=100.0, step=10.0, format="%.2f")
                
        st.markdown("---")
        
        if "Yüzdelik" in hesap_yontemi:
            toplam_kesinti_orani = (in_komisyon + in_hedef_val) / 100.0
            if toplam_kesinti_orani >= 1.0:
                st.error("❌ Hata: Komisyon ve Kâr Marjı toplamı %100 veya daha fazla olamaz! Lütfen oranları düşürün.")
                ideal_fiyat = 0.0
            else:
                ideal_fiyat = (in_maliyet + in_kargo) / (1.0 - toplam_kesinti_orani)
        else:
            kom_orani = in_komisyon / 100.0
            if kom_orani >= 1.0:
                st.error("❌ Hata: Komisyon oranı %100 olamaz!")
                ideal_fiyat = 0.0
            else:
                ideal_fiyat = (in_maliyet + in_kargo + in_hedef_val) / (1.0 - kom_orani)
                
        if ideal_fiyat > 0:
            komisyon_tl = ideal_fiyat * (in_komisyon / 100.0)
            toplam_gider = in_maliyet + in_kargo + komisyon_tl
            net_kar = ideal_fiyat - toplam_gider
            kar_marji = (net_kar / ideal_fiyat * 100.0) if ideal_fiyat > 0 else 0.0
            
            st.markdown("### 💡 Önerilen Satış Fiyatı ve Gider Kırılımı")
            r1, r2, r3, r4 = st.columns(4)
            with r1: st.markdown(f'<div class="calc-metric" style="border-left-color:#8E44AD;"><b>İdeal Satış Fiyatı:</b><br><span style="font-size:26px; font-weight:800; color:#8E44AD;">{ideal_fiyat:,.2f} TL</span></div>', unsafe_allow_html=True)
            with r2: st.markdown(f'<div class="calc-metric" style="border-left-color:#2ECC71;"><b>Net Kâr Tutarı:</b><br><span style="font-size:26px; font-weight:800; color:#2ECC71;">{net_kar:,.2f} TL</span></div>', unsafe_allow_html=True)
            with r3: st.markdown(f'<div class="calc-metric" style="border-left-color:#2980B9;"><b>Gerçek Kâr Marjı:</b><br><span style="font-size:26px; font-weight:800; color:#2980B9;">% {kar_marji:.2f}</span></div>', unsafe_allow_html=True)
            with r4: st.markdown(f'<div class="calc-metric" style="border-left-color:#E74C3C;"><b>Komisyon Kesintisi:</b><br><span style="font-size:26px; font-weight:800; color:#E74C3C;">{komisyon_tl:,.2f} TL</span></div>', unsafe_allow_html=True)
            
            st.write("")
            with st.expander("📊 Satış Fiyatı Dağılımı (Cironuz Nereye Gidiyor?)", expanded=True):
                p1, p2, p3, p4 = st.columns(4)
                with p1: st.info(f"📦 **Maliyet Payı:** {in_maliyet:,.2f} TL (% {(in_maliyet/ideal_fiyat*100):.1f})")
                with p2: st.warning(f"🚚 **Kargo Payı:** {in_kargo:,.2f} TL (% {(in_kargo/ideal_fiyat*100):.1f})")
                with p3: st.error(f"🤝 **Komisyon Payı:** {komisyon_tl:,.2f} TL (% {(komisyon_tl/ideal_fiyat*100):.1f})")
                with p4: st.success(f"💰 **Size Kalan Kâr:** {net_kar:,.2f} TL (% {(net_kar/ideal_fiyat*100):.1f})")

    with tab_toplu:
        st.markdown("#### 📦 Veritabanındaki Tüm Ürünleri Tek Tuşla Fiyatlandırın")
        st.write("Sistemdeki `Maliyet Yönetimi` tablosunda kayıtlı olan tüm ürünleriniz için istediğiniz kâr marjını uygulayarak önerilen satış fiyatlarını otomatik oluşturur.")
        
        db = load_db()
        if db.empty:
            st.warning("⚠ Veritabanında kayıtlı ürün bulunmuyor. Önce sol menüdeki '📦 Maliyet Yönetimi' sayfasından ürün ekleyin.")
        else:
            col_b1, col_b2 = st.columns(2)
            with col_b1:
                toplu_yontem = st.selectbox("Toplu Hesaplama Hedefi:", ["📊 Sabit Kâr Marjı (%) Uygula", "💵 Sabit Net Kâr Tutarı (TL) Uygula"])
            with col_b2:
                if "Marjı" in toplu_yontem:
                    toplu_hedef_val = st.number_input("Tüm Ürünler İçin Hedef Kâr Marjı (%)", min_value=1.0, max_value=80.0, value=30.0, step=1.0)
                else:
                    toplu_hedef_val = st.number_input("Tüm Ürünler İçin Hedef Net Kâr (TL)", min_value=1.0, value=120.0, step=10.0)
                    
            if st.button("🚀 Tüm Ürünler İçin İdeal Fiyatları Hesapla", type="primary", use_container_width=True):
                with st.spinner("Ürünler fiyatlandırılıyor..."):
                    df_res = db.copy()
                    ideal_fiyatlar, komisyonlar, net_karlar, marjlar = [], [], [], []
                    
                    for idx, row in df_res.iterrows():
                        m_val, k_val, kom_val = row['Maliyet (TL)'], row['Kargo (TL)'], row['Komisyon (%)']
                        
                        if "Marjı" in toplu_yontem:
                            t_rate = (kom_val + toplu_hedef_val) / 100.0
                            if t_rate >= 1.0:
                                i_fiyat = 0.0
                            else:
                                i_fiyat = (m_val + k_val) / (1.0 - t_rate)
                        else:
                            k_rate = kom_val / 100.0
                            if k_rate >= 1.0:
                                i_fiyat = 0.0
                            else:
                                i_fiyat = (m_val + k_val + toplu_hedef_val) / (1.0 - k_rate)
                                
                        kom_tl = i_fiyat * (kom_val / 100.0) if i_fiyat > 0 else 0.0
                        n_kar = i_fiyat - (m_val + k_val + kom_tl) if i_fiyat > 0 else 0.0
                        k_marji = (n_kar / i_fiyat * 100.0) if i_fiyat > 0 else 0.0
                        
                        ideal_fiyatlar.append(i_fiyat)
                        komisyonlar.append(kom_tl)
                        net_karlar.append(n_kar)
                        marjlar.append(k_marji)
                        
                    df_res['Önerilen Satış Fiyatı (TL)'] = ideal_fiyatlar
                    df_res['Komisyon Tutarı (TL)'] = komisyonlar
                    df_res['Net Kâr (TL)'] = net_karlar
                    df_res['Kâr Marjı (%)'] = marjlar
                    
                    st.success("✅ Toplu fiyatlandırma başarıyla tamamlandı!")
                    st.dataframe(tablayi_1den_baslat(df_res).style.format({
                        'Maliyet (TL)': '{:,.2f} TL', 'Kargo (TL)': '{:,.2f} TL', 'Komisyon (%)': '% {:.2f}',
                        'Önerilen Satış Fiyatı (TL)': '{:,.2f} TL', 'Komisyon Tutarı (TL)': '{:,.2f} TL',
                        'Net Kâr (TL)': '{:,.2f} TL', 'Kâr Marjı (%)': '% {:.2f}'
                    }), use_container_width=True)
                    
                    out_exc = BytesIO()
                    with pd.ExcelWriter(out_exc, engine='openpyxl') as wr:
                        tablayi_1den_baslat(df_res).reset_index().to_excel(wr, index=False, sheet_name='İdeal Fiyat Listesi')
                        wb = wr.book; ws = wb.active
                        fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
                        font = Font(bold=True, color="FFFFFF")
                        for col_idx, cell in enumerate(ws[1], 1): cell.fill = fill; cell.font = font
                        for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width = 18
                    out_exc.seek(0)
                    
                    st.download_button(
                        label="📥 İdeal Fiyat Listesini Excel Olarak İndir",
                        data=out_exc,
                        file_name=f"Toplu_Ideal_Fiyat_Listesi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

# ==========================================
# SAYFA 3: TRENDYOL KAMPANYA ANALİZİ
# ==========================================
elif menu == "🚀 Trendyol Yıldızlı Fiyat":
    st.markdown('<div class="main-title">📈 Trendyol Yıldızlı Ürün Analizi</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">Trendyol\'dan indirdiğiniz "Yıldızlı Ürün Etiketleri" dosyasını yükleyin. Sistem 3 Yıldız > 2 Yıldız > 1 Yıldız sırasıyla kârı test eder.</div>', unsafe_allow_html=True)
    
    db = load_db()
    if db.empty:
        st.error("❌ Lütfen önce sol menüden ürün maliyetlerinizi girin!")
        st.stop()
        
    st.markdown("### 🎯 Kârlılık Kriterleri")
    c1, c2 = st.columns(2)
    with c1: min_kar_marji = st.number_input("Minimum Hedef Kâr Marjı (%)", min_value=-50.0, value=35.0, step=1.0, key="ty_marj")
    with c2: min_net_kar_tl = st.number_input("Minimum Net Kâr Tutarı (TL)", min_value=0.0, value=100.0, step=1.0, key="ty_tl")
    st.markdown("---")
    
    kampanya_file = st.file_uploader("Trendyol 'Yıldızlı Ürün Etiketleri' Dosyasını Yükleyin", type=['xlsx', 'csv'], key="ty_file")
    
    if kampanya_file:
        orijinal_dosya_ismi = kampanya_file.name
        if orijinal_dosya_ismi.endswith('.csv'):
            try: df_kampanya = pd.read_csv(kampanya_file, sep=None, engine='python')
            except: kampanya_file.seek(0); df_kampanya = pd.read_csv(kampanya_file, delimiter=';')
        else: df_kampanya = pd.read_excel(kampanya_file)
            
        cols = list(df_kampanya.columns)
        barkod_col = next((c for c in cols if 'BARKOD' in c.upper()), cols[1] if len(cols)>1 else cols[0])
        fiyat_1_yildiz = next((c for c in cols if '1 YILDIZ ÜST FİYAT' in c.upper()), None)
        fiyat_2_yildiz = next((c for c in cols if '2 YILDIZ ÜST FİYAT' in c.upper()), None)
        fiyat_3_yildiz = next((c for c in cols if '3 YILDIZ ÜST FİYAT' in c.upper()), None)
        yeni_fiyat_col = next((c for c in cols if 'YENİ TSF' in c.upper()), None)
        
        if st.button("⚡ Otomatik Fiyatlandır (Trendyol)", use_container_width=True):
            if not all([fiyat_1_yildiz, fiyat_2_yildiz, fiyat_3_yildiz, yeni_fiyat_col]):
                st.error("❌ Yüklediğiniz dosyada Yıldızlı Fiyat sütunları bulunamadı. Doğru şablonu yüklediğinizden emin olun.")
            else:
                with st.spinner("⏳ Fiyatlandırma senaryoları test ediliyor..."):
                    islem_df = df_kampanya.copy()
                    db_copy = db.copy()
                    islem_df['_kamp_barkod'] = islem_df[barkod_col].astype(str).str.strip()
                    db_copy['_db_barkod'] = db_copy['Barkod'].astype(str).str.strip()
                    merge_df = pd.merge(islem_df, db_copy, left_on='_kamp_barkod', right_on='_db_barkod', how='left')
                    
                    for col in [fiyat_1_yildiz, fiyat_2_yildiz, fiyat_3_yildiz]:
                        merge_df[col + '_num'] = merge_df[col].apply(temizle_ve_sayiya_donustur)
                        
                    secilen_fiyatlar, secilen_yildizlar, hesaplanan_karlar, hesaplanan_marjlar = [], [], [], []
                    test_siralamasi = [("3 Yıldız", fiyat_3_yildiz+'_num'), ("2 Yıldız", fiyat_2_yildiz+'_num'), ("1 Yıldız", fiyat_1_yildiz+'_num')]
                    
                    for idx, row in merge_df.iterrows():
                        if pd.isna(row['_db_barkod']):
                            secilen_fiyatlar.append(np.nan); secilen_yildizlar.append("Sistemde Yok"); hesaplanan_karlar.append(0); hesaplanan_marjlar.append(0); continue
                        maliyet, kargo, kom_orani = row['Maliyet (TL)'], row['Kargo (TL)'], row['Komisyon (%)']
                        uygun_fiyat, secili_yildiz, net_kar, kar_marji = np.nan, "Elenmiş", 0, 0
                        
                        for yildiz_isim, f_col in test_siralamasi:
                            fiyat = row[f_col]
                            if fiyat <= 0: continue
                            komisyon_tl = fiyat * (kom_orani / 100)
                            n_kar = fiyat - (maliyet + kargo + komisyon_tl)
                            k_marji = (n_kar / fiyat) * 100 if fiyat > 0 else 0
                            if n_kar >= min_net_kar_tl and k_marji >= min_kar_marji:
                                uygun_fiyat, secili_yildiz, net_kar, kar_marji = fiyat, yildiz_isim, n_kar, k_marji; break 
                                
                        secilen_fiyatlar.append(uygun_fiyat); secilen_yildizlar.append(secili_yildiz); hesaplanan_karlar.append(net_kar); hesaplanan_marjlar.append(kar_marji)
                        
                    islem_df['Seçilen Yıldız'] = secilen_yildizlar; islem_df['Net Kâr (TL)'] = hesaplanan_karlar; islem_df['Kâr Marjı (%)'] = hesaplanan_marjlar
                    
                    def format_fiyat(val):
                        if pd.isna(val) or val == 0: return ""
                        return str(round(val, 2)).replace('.', ',')
                        
                    islem_df[yeni_fiyat_col] = [format_fiyat(f) for f in secilen_fiyatlar]
                    basarili_df = islem_df[islem_df['Seçilen Yıldız'].isin(["1 Yıldız", "2 Yıldız", "3 Yıldız"])].copy()
                    elenen_df = islem_df[islem_df['Seçilen Yıldız'] == "Elenmiş"].copy()
                    db_yok_df = islem_df[islem_df['Seçilen Yıldız'] == "Sistemde Yok"].copy()
                    
                    st.markdown("### 📊 Analiz Özeti")
                    m1, m2, m3, m4 = st.columns(4)
                    with m1: st.markdown(f'<div class="metric-box" style="border-color:#3498DB;"><b>İncelenen Ürün:</b><br><span style="font-size:24px; font-weight:bold;">{len(islem_df)}</span></div>', unsafe_allow_html=True)
                    with m2: st.markdown(f'<div class="metric-box" style="border-color:#2ECC71;"><b>Fiyat Atanan:</b><br><span style="font-size:24px; font-weight:bold; color:#2ECC71;">{len(basarili_df)}</span></div>', unsafe_allow_html=True)
                    with m3: st.markdown(f'<div class="metric-box" style="border-color:#E74C3C;"><b>Zarar Sebebiyle Elenen:</b><br><span style="font-size:24px; font-weight:bold; color:#E74C3C;">{len(elenen_df)}</span></div>', unsafe_allow_html=True)
                    with m4: st.markdown(f'<div class="metric-box" style="border-color:#F39C12;"><b>Maliyeti Girilmeyen:</b><br><span style="font-size:24px; font-weight:bold; color:#F39C12;">{len(db_yok_df)}</span></div>', unsafe_allow_html=True)
                    
                    st.write("#### 🎯 Fiyat Ataması Yapılan Ürünler (Önizleme)")
                    if len(basarili_df) > 0:
                        orijinal_fiyat_col = next((c for c in cols if 'TRENDYOL SATIŞ FİYATI' in c.upper() or 'SATIŞ FİYATI' in c.upper()), None)
                        goster_cols = [barkod_col]
                        if orijinal_fiyat_col: goster_cols.append(orijinal_fiyat_col)
                        goster_cols.extend([yeni_fiyat_col, 'Seçilen Yıldız', 'Net Kâr (TL)', 'Kâr Marjı (%)'])
                        st.dataframe(tablayi_1den_baslat(basarili_df[goster_cols]).style.format({'Net Kâr (TL)': '{:.2f} TL', 'Kâr Marjı (%)': '% {:.2f}'}), use_container_width=True)
                    else: st.warning("Hiçbir ürün kriterleri karşılamadı.")
                    
                    output = BytesIO(); export_df = islem_df[cols].copy() 
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        export_df.to_excel(writer, index=False, sheet_name='Sheet1'); workbook = writer.book; worksheet = workbook.active
                        header_fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid"); header_font = Font(bold=True, color="FFFFFF")
                        p_col_idx = export_df.columns.get_loc(yeni_fiyat_col) + 1
                        for col_idx, col_name in enumerate(export_df.columns, 1):
                            cell = worksheet.cell(row=1, column=col_idx); cell.fill = header_fill; cell.font = header_font
                            if col_idx == p_col_idx: cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                        for col in worksheet.columns: worksheet.column_dimensions[get_column_letter(col[0].column)].width = 15
                    output.seek(0)
                    st.success("✅ Excel dosyanız yüklendiği orijinal isimle indirilmeye hazır!")
                    st.download_button(label="📥 Trendyol İçin Hazır Excel'i İndir", data=output, file_name=orijinal_dosya_ismi.rsplit('.', 1)[0] + ".xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

# ==========================================
# SAYFA 4: HEPSİBURADA KAMPANYA ANALİZİ
# ==========================================
elif menu == "💜 Hepsiburada Avantajlı Teklif":
    st.markdown('<div class="hb-title">💜 Hepsiburada Kampanya Analizi</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">Hepsiburada\'dan indirdiğiniz "Listelerim" dosyasını yükleyin. Sepet kampanyaları veya standart fiyat kampanyaları için uygun kârlılığı otomatik hesaplar.</div>', unsafe_allow_html=True)
    
    db = load_db()
    if db.empty: st.error("❌ Lütfen önce sol menüden ürün maliyetlerinizi girin!"); st.stop()
        
    st.markdown("### 🎯 Kârlılık Kriterleri")
    c1, c2 = st.columns(2)
    with c1: min_kar_marji = st.number_input("Minimum Hedef Kâr Marjı (%)", min_value=-50.0, value=35.0, step=1.0, key="hb_marj")
    with c2: min_net_kar_tl = st.number_input("Minimum Net Kâr Tutarı (TL)", min_value=0.0, value=100.0, step=1.0, key="hb_tl")
    st.markdown("---")
    
    kampanya_file = st.file_uploader("Hepsiburada 'Listelerim' Dosyasını Yükleyin", type=['xlsx', 'csv'], key="hb_file")
    
    if kampanya_file:
        orijinal_dosya_ismi = kampanya_file.name
        if orijinal_dosya_ismi.endswith('.csv'):
            try: df_kampanya = pd.read_csv(kampanya_file, sep=None, engine='python')
            except: kampanya_file.seek(0); df_kampanya = pd.read_csv(kampanya_file, delimiter=';')
        else: df_kampanya = pd.read_excel(kampanya_file)
            
        cols = list(df_kampanya.columns)
        st.write("#### ⚙ Sütun Eşleştirme (Hepsiburada Formatı İçin)")
        
        col1, col2, col3, col4 = st.columns(4)
        with col1: barkod_col = st.selectbox("Barkod / SKU Sütunu", cols, index=find_default_col(cols, ["barkod", "barcode", "sku", "stok", "merchant"]))
        with col2: eski_fiyat_col = st.selectbox("Mevcut Satış Fiyatı", cols, index=find_default_col(cols, ["satış", "satis", "psf"], exclude_keywords=["kampanya", "teklif", "max"]))
        with col3:
            is_normal_campaign = st.checkbox("🎯 Standart Kampanya (HB Max Fiyat Kuralı)")
            if not is_normal_campaign:
                sepet_indirimi = st.number_input("🛒 Kampanya İndirim Oranı (%) (Örn: %15 için 15 yazın)", min_value=1.0, max_value=99.0, value=15.0, step=1.0)
            else:
                sepet_indirimi = 0.0
            max_fiyat_col = st.selectbox("Girebileceğiniz Max. Fiyat", cols, index=find_default_col(cols, ["max", "maksimum", "girebileceğiniz"]))
        with col4: kampanya_fiyat_col = st.selectbox("Hepsiburada Paneline Yüklenecek Fiyat", cols, index=find_default_col(cols, ["uygulanacağı", "kampanya", "önerilen", "teklif", "avantajlı"], exclude_keywords=["durum"]))
        
        if st.button("⚡ Otomatik Fiyatlandır (Hepsiburada)", use_container_width=True):
            with st.spinner("⏳ Hepsiburada teklifleri kârlılık testinden geçiriliyor..."):
                islem_df = df_kampanya.copy()
                db_copy = db.copy()
                islem_df['_kamp_barkod'] = islem_df[barkod_col].astype(str).str.strip()
                db_copy['_db_barkod'] = db_copy['Barkod'].astype(str).str.strip()
                merge_df = pd.merge(islem_df, db_copy, left_on='_kamp_barkod', right_on='_db_barkod', how='left')
                
                durum_list, hesaplanan_karlar, hesaplanan_marjlar, katilim_fiyati = [], [], [], []
                for idx, row in merge_df.iterrows():
                    if pd.isna(row['_db_barkod']):
                        durum_list.append("Sistemde Yok"); hesaplanan_karlar.append(0); hesaplanan_marjlar.append(0); katilim_fiyati.append(np.nan); continue
                    maliyet, kargo, kom_orani = row['Maliyet (TL)'], row['Kargo (TL)'], row['Komisyon (%)']
                    fiyat = temizle_ve_sayiya_donustur(row[eski_fiyat_col])
                    if sepet_indirimi > 0: fiyat = fiyat * (1 - (sepet_indirimi / 100))
                    if fiyat <= 0:
                        durum_list.append("Elenmiş"); hesaplanan_karlar.append(0); hesaplanan_marjlar.append(0); katilim_fiyati.append(np.nan); continue
                    komisyon_tl = fiyat * (kom_orani / 100)
                    n_kar = fiyat - (maliyet + kargo + komisyon_tl)
                    k_marji = (n_kar / fiyat) * 100 if fiyat > 0 else 0
                    if n_kar >= min_net_kar_tl and k_marji >= min_kar_marji:
                        durum_list.append("Kabul Edildi"); hesaplanan_karlar.append(n_kar); hesaplanan_marjlar.append(k_marji); katilim_fiyati.append(np.nan if is_normal_campaign else fiyat)
                    else:
                        durum_list.append("Elenmiş"); hesaplanan_karlar.append(n_kar); hesaplanan_marjlar.append(k_marji); katilim_fiyati.append(np.nan)
                        
                islem_df['Kampanya Durumu'] = durum_list; islem_df['Net Kâr (TL)'] = hesaplanan_karlar; islem_df['Kâr Marjı (%)'] = hesaplanan_marjlar
                
                def hb_format(val):
                    if pd.isna(val) or val == 0: return ""
                    return str(round(val, 2)).replace('.', ',')
                islem_df[kampanya_fiyat_col] = [hb_format(f) for f in katilim_fiyati]
                
                basarili_df = islem_df[islem_df['Kampanya Durumu'] == "Kabul Edildi"].copy()
                elenen_df = islem_df[islem_df['Kampanya Durumu'] == "Elenmiş"].copy()
                db_yok_df = islem_df[islem_df['Kampanya Durumu'] == "Sistemde Yok"].copy()
                
                st.markdown("### 📊 Hepsiburada Analiz Özeti")
                m1, m2, m3, m4 = st.columns(4)
                with m1: st.markdown(f'<div class="hb-metric" style="border-color:#3498DB;"><b>İncelenen Ürün:</b><br><span style="font-size:24px; font-weight:bold;">{len(islem_df)}</span></div>', unsafe_allow_html=True)
                with m2: st.markdown(f'<div class="hb-metric" style="border-color:#2ECC71;"><b>Teklifi Uygun Olan:</b><br><span style="font-size:24px; font-weight:bold; color:#2ECC71;">{len(basarili_df)}</span></div>', unsafe_allow_html=True)
                with m3: st.markdown(f'<div class="hb-metric" style="border-color:#E74C3C;"><b>Zarar/Düşük Kâr:</b><br><span style="font-size:24px; font-weight:bold; color:#E74C3C;">{len(elenen_df)}</span></div>', unsafe_allow_html=True)
                with m4: st.markdown(f'<div class="hb-metric" style="border-color:#F39C12;"><b>Maliyeti Girilmeyen:</b><br><span style="font-size:24px; font-weight:bold; color:#F39C12;">{len(db_yok_df)}</span></div>', unsafe_allow_html=True)
                
                st.write("#### 🎯 Hepsiburada Kampanyasına Katılacak Ürünler")
                if len(basarili_df) > 0:
                    st.dataframe(tablayi_1den_baslat(basarili_df[[barkod_col, eski_fiyat_col, 'Net Kâr (TL)', 'Kâr Marjı (%)']]).style.format({'Net Kâr (TL)': '{:.2f} TL', 'Kâr Marjı (%)': '% {:.2f}'}), use_container_width=True)
                else: st.warning("Hedef kâr marjınızı karşılayan hiçbir HB teklifi bulunamadı.")
                
                output = BytesIO(); export_df = islem_df[cols].copy() 
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    if len(basarili_df) > 0: basarili_df[cols].to_excel(writer, index=False, sheet_name='Uygun Teklifler')
                    islem_df[cols + ['Kampanya Durumu', 'Net Kâr (TL)', 'Kâr Marjı (%)']].to_excel(writer, index=False, sheet_name='Tüm Analiz Sonucu')
                    workbook = writer.book
                    for sheetname in workbook.sheetnames:
                        worksheet = workbook[sheetname]; header_fill = PatternFill(start_color="FF6700", end_color="FF6700", fill_type="solid"); header_font = Font(bold=True, color="FFFFFF")
                        if sheetname == 'Uygun Teklifler':
                            h_col_idx = export_df.columns.get_loc(kampanya_fiyat_col) + 1
                            for col_idx, col_name in enumerate(worksheet[1], 1):
                                col_name.fill = header_fill; col_name.font = header_font
                                if col_idx == h_col_idx and sepet_indirimi > 0: col_name.fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
                        else:
                            for col_idx, col_name in enumerate(worksheet[1], 1): col_name.fill = header_fill; col_name.font = header_font
                        for col in worksheet.columns: worksheet.column_dimensions[get_column_letter(col[0].column)].width = 15
                output.seek(0)
                st.success("✅ Hepsiburada dosyanız hazır! İndirdiğiniz dosyadaki 'Uygun Teklifler' sekmesini doğrudan panelinize yükleyebilirsiniz.")
                st.download_button(label="📥 Hepsiburada İçin Hazır Excel'i İndir", data=output, file_name="HB_Kampanya_Sonucu.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

# ==========================================
# SAYFA 5: TRENDYOL DETAYLI SATIŞ ANALİZİ (API) - CANLI DÜZELTİLDİ
# ==========================================
elif menu == "📊 Trendyol Satış Analizi (API)":
    st.markdown('<div class="sales-title">📊 Trendyol Detaylı Satış ve Kârlılık Analizi (API)</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">Anlık, Günlük, Haftalık, Aylık veya Yıllık tüm satışlarınızı doğrudan API ile çekin. Canlı satışlar artık eksiksiz çekiliyor.</div>', unsafe_allow_html=True)
    
    db = load_db()
    if db.empty: st.error("❌ Veritabanı boş! Masrafları hesaplayabilmek için önce 'Maliyet Yönetimi' sayfasından maliyetlerinizi kaydetmelisiniz."); st.stop()
    
    api = load_api_settings()
    if not (api["ty_seller_id"] and api["ty_api_key"] and api["ty_api_secret"]):
        st.warning("⚠ Trendyol API bilgileri sistemde kayıtlı değil! Lütfen sol menüden **'⚙ Ayarlar & API'** sekmesine giderek bilgilerinizi bir kez kaydedin.")
        st.stop()
    
    now = datetime.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    st.markdown("### 🗓 Zaman ve Tarih Filtresi Seçimi")
    zaman_filtresi = st.radio("İncelemek İstediğiniz Dönemi Seçin:", [
        "⚡ Anlık / Bugün (Bugünün Satışları)",
        "📅 Günlük / Dün (Dünün Satışları)",
        "🗓 Haftalık (Son 7 Gün)",
        "📆 Aylık (Son 30 Gün)",
        "🚀 Yıllık (Bu Yıl - 1 Ocak'tan Bugüne)",
        "🔍 İki Tarih Arası (Özel Seçim)"
    ], horizontal=True)
    
    if "Anlık / Bugün" in zaman_filtresi:
        start_dt = today_start
        end_dt = now
    elif "Günlük / Dün" in zaman_filtresi:
        start_dt = today_start - timedelta(days=1)
        end_dt = today_start - timedelta(seconds=1)
    elif "Haftalık" in zaman_filtresi:
        start_dt = today_start - timedelta(days=7)
        end_dt = now
    elif "Aylık" in zaman_filtresi:
        start_dt = today_start - timedelta(days=30)
        end_dt = now
    elif "Yıllık" in zaman_filtresi:
        start_dt = datetime(now.year, 1, 1, 0, 0, 0)
        end_dt = now
    elif "İki Tarih Arası" in zaman_filtresi:
        col_d1, col_d2 = st.columns(2)
        with col_d1: t_basla = st.date_input("Başlangıç Tarihi", value=(now - timedelta(days=15)).date())
        with col_d2: t_bitis = st.date_input("Bitiş Tarihi", value=now.date())
        start_dt = datetime.combine(t_basla, datetime.min.time())
        end_dt = datetime.combine(t_bitis, datetime.max.time())
        
    st.info(f"📍 **Seçilen Analiz Aralığı:** {turkce_tarih_format(start_dt)} - {turkce_tarih_format(end_dt)}")
    
    if st.button("🔄 Siparişleri Çek ve Detaylı Analiz Et", type="primary", use_container_width=True):
        with st.spinner("Trendyol API'den siparişler çekiliyor, maliyet ve kargo kesintileri hesaplanıyor..."):
            orders = fetch_ty_orders(start_dt, end_dt)
            
            if orders:
                db_copy = db.copy()
                db_copy['_db_barkod'] = db_copy['Barkod'].astype(str).str.strip()
                siparis_detaylari = []
                
                for order in orders:
                    order_ts = order.get("orderDate", 0) / 1000
                    try:
                        order_dt = datetime.fromtimestamp(order_ts)
                    except:
                        continue
                    
                    # API zaten tarih aralığını filtreledi, canlı veriyi kaçırmamak için ikinci filtre kaldırıldı
                    sip_tarihi_str = turkce_tarih_format(order_dt)
                    siparis_no = order.get("orderNumber", "")
                    
                    for line in order.get("lines", []):
                        barkod = str(line.get("barcode", "")).strip()
                        urun_adi = line.get("productName", "")[:50]
                        adet = int(line.get("quantity", 1))
                        fiyat = float(line.get("price", 0.0) or line.get("amount", 0.0) or 0.0)
                        ciro = adet * fiyat
                        
                        eslesme = db_copy[db_copy['_db_barkod'] == barkod]
                        if len(eslesme) > 0:
                            m_tl = eslesme.iloc[0]['Maliyet (TL)'] * adet
                            kg_tl = eslesme.iloc[0]['Kargo (TL)'] * adet
                            kom_orani = eslesme.iloc[0]['Komisyon (%)']
                            kom_tl = ciro * (kom_orani / 100)
                        else:
                            m_tl, kg_tl, kom_tl, kom_orani = 0.0, 0.0, 0.0, 0.0
                            
                        net_k = ciro - (m_tl + kg_tl + kom_tl)
                        marj = (net_k / ciro * 100) if ciro > 0 else 0.0
                        
                        siparis_detaylari.append({
                            "Tarih": sip_tarihi_str, "Sipariş No": siparis_no, "Barkod": barkod, "Ürün Adı": urun_adi,
                            "Adet": adet, "Birim Fiyat": fiyat, "Satış Tutarı (Ciro)": ciro,
                            "Maliyet (TL)": m_tl, "Kargo (TL)": kg_tl, "Komisyon (%)": kom_orani,
                            "Komisyon (TL)": kom_tl, "Net Kâr (TL)": net_k, "Kâr Marjı (%)": marj
                        })
                        
                if siparis_detaylari:
                    df_sip = pd.DataFrame(siparis_detaylari)
                    st.session_state["ty_satis_raporu"] = df_sip
                    st.success(f"✅ {len(orders)} sipariş, {len(df_sip)} satır başarıyla çekildi!")
                else:
                    st.warning("⚠ Seçilen tarih aralığında sipariş kaydı bulunamadı.")
                    st.session_state["ty_satis_raporu"] = pd.DataFrame()
            else:
                st.warning("⚠ Trendyol sunucularından sipariş verisi alınamadı veya seçilen tarihte satış yok.")
                st.session_state["ty_satis_raporu"] = pd.DataFrame()
            
    if "ty_satis_raporu" in st.session_state and not st.session_state["ty_satis_raporu"].empty:
        df_sip = st.session_state["ty_satis_raporu"]
        
        top_adet = df_sip["Adet"].sum()
        top_ciro = df_sip["Satış Tutarı (Ciro)"].sum()
        top_maliyet = df_sip["Maliyet (TL)"].sum()
        top_kargo = df_sip["Kargo (TL)"].sum()
        top_komisyon = df_sip["Komisyon (TL)"].sum()
        top_masraf = top_maliyet + top_kargo + top_komisyon
        top_kar = df_sip["Net Kâr (TL)"].sum()
        ort_marj = (top_kar / top_ciro * 100) if top_ciro > 0 else 0.0
        
        st.markdown("---")
        st.markdown("### 📊 Seçilen Dönem Performans Özeti")
        k1, k2, k3, k4, k5 = st.columns(5)
        with k1: st.markdown(f'<div class="sales-metric"><b>Satış Adedi:</b><br><span style="font-size:22px; font-weight:bold; color:#2980B9;">{int(top_adet):,} Adet</span></div>', unsafe_allow_html=True)
        with k2: st.markdown(f'<div class="sales-metric"><b>Satış Tutarı (Ciro):</b><br><span style="font-size:22px; font-weight:bold; color:#27AE60;">{top_ciro:,.2f} TL</span></div>', unsafe_allow_html=True)
        with k3: st.markdown(f'<div class="sales-metric"><b>Toplam Masraflar:</b><br><span style="font-size:22px; font-weight:bold; color:#E74C3C;">{top_masraf:,.2f} TL</span><br><small style="color:#7f8c8d;">Maliyet+Kargo+Kom.</small></div>', unsafe_allow_html=True)
        with k4: st.markdown(f'<div class="sales-metric" style="border-left-color: {"#2ECC71" if top_kar>=0 else "#E74C3C"};"><b>Net Kâr:</b><br><span style="font-size:22px; font-weight:bold; color:{"#2ECC71" if top_kar>=0 else "#E74C3C"};">{top_kar:,.2f} TL</span></div>', unsafe_allow_html=True)
        with k5: st.markdown(f'<div class="sales-metric"><b>Ort. Kâr Marjı:</b><br><span style="font-size:22px; font-weight:bold; color:#8E44AD;">% {ort_marj:.2f}</span></div>', unsafe_allow_html=True)
        
        with st.expander("ℹ Masraf Kırılımı Detayını Göster (Maliyet, Kargo ve Komisyon Özetleri)"):
            mc1, mc2, mc3 = st.columns(3)
            with mc1: st.info(f"📦 **Ürün Maliyeti Toplamı:** {top_maliyet:,.2f} TL (% {(top_maliyet/top_ciro*100) if top_ciro>0 else 0:.1f})")
            with mc2: st.warning(f"🚚 **Toplam Kargo Gideri:** {top_kargo:,.2f} TL (% {(top_kargo/top_ciro*100) if top_ciro>0 else 0:.1f})")
            with mc3: st.error(f"🤝 **Trendyol Komisyon Kesintisi:** {top_komisyon:,.2f} TL (% {(top_komisyon/top_ciro*100) if top_ciro>0 else 0:.1f})")

        st.markdown("### 🏆 Dönemin Şampiyon Ürünleri")
        urun_bazli = df_sip.groupby('Barkod').agg({
            'Ürün Adı': 'first',
            'Adet': 'sum',
            'Satış Tutarı (Ciro)': 'sum',
            'Net Kâr (TL)': 'sum'
        }).reset_index()
        
        en_cok_satan = urun_bazli.sort_values(by='Adet', ascending=False).iloc[0] if len(urun_bazli)>0 else None
        en_yuksek_ciro = urun_bazli.sort_values(by='Satış Tutarı (Ciro)', ascending=False).iloc[0] if len(urun_bazli)>0 else None
        en_cok_kar = urun_bazli.sort_values(by='Net Kâr (TL)', ascending=False).iloc[0] if len(urun_bazli)>0 else None
        
        h1, h2, h3 = st.columns(3)
        with h1:
            if en_cok_satan is not None: st.markdown(f'<div class="highlight-card">👑 <b>En Çok Satılan Ürün</b><br><span style="color:#2980B9; font-weight:bold; font-size:15px;">{str(en_cok_satan["Ürün Adı"])[:38]}...</span><br><br><span style="font-size:20px; font-weight:800; color:#2C3E50;">{int(en_cok_satan["Adet"]):,} Adet</span> Satış</div>', unsafe_allow_html=True)
        with h2:
            if en_yuksek_ciro is not None: st.markdown(f'<div class="highlight-card">💎 <b>En Çok Ciro Getiren Ürün</b><br><span style="color:#27AE60; font-weight:bold; font-size:15px;">{str(en_yuksek_ciro["Ürün Adı"])[:38]}...</span><br><br><span style="font-size:20px; font-weight:800; color:#2C3E50;">{en_yuksek_ciro["Satış Tutarı (Ciro)"]:,.2f} TL</span> Ciro</div>', unsafe_allow_html=True)
        with h3:
            if en_cok_kar is not None: st.markdown(f'<div class="highlight-card">🚀 <b>En Çok Kâr Bırakan Ürün</b><br><span style="color:#8E44AD; font-weight:bold; font-size:15px;">{str(en_cok_kar["Ürün Adı"])[:38]}...</span><br><br><span style="font-size:20px; font-weight:800; color:#2ECC71;">{en_cok_kar["Net Kâr (TL)"]:,.2f} TL</span> Net Kâr</div>', unsafe_allow_html=True)

        st.markdown("---")
        
        tab_ozet, tab_siparis = st.tabs(["📦 Hangi Üründen Kaç Tane Satılmış? (Ürün Bazlı Kâr Analizi)", "📜 Satır Satır Detaylı Sipariş ve Kesinti Listesi"])
        
        with tab_ozet:
            st.write("Dönem boyunca hangi üründen toplam kaç adet satıldığını, ne kadar ciro ve kâr bıraktığını detaylı inceleyin (Sıralama: Çok Satandan Az Satana):")
            ozet_t = df_sip.groupby('Barkod').agg({
                'Ürün Adı': 'first',
                'Adet': 'sum',
                'Satış Tutarı (Ciro)': 'sum',
                'Maliyet (TL)': 'sum',
                'Kargo (TL)': 'sum',
                'Komisyon (TL)': 'sum',
                'Net Kâr (TL)': 'sum'
            }).reset_index()
            
            ozet_t['Kâr Marjı (%)'] = np.where(ozet_t['Satış Tutarı (Ciro)']>0, (ozet_t['Net Kâr (TL)'] / ozet_t['Satış Tutarı (Ciro)'] * 100), 0.0)
            ozet_t = ozet_t.sort_values(by='Adet', ascending=False)
            
            st.dataframe(tablayi_1den_baslat(ozet_t).style.format({
                'Adet': '{:,.0f}', 'Satış Tutarı (Ciro)': '{:,.2f} TL', 'Maliyet (TL)': '{:,.2f} TL',
                'Kargo (TL)': '{:,.2f} TL', 'Komisyon (TL)': '{:,.2f} TL', 'Net Kâr (TL)': '{:,.2f} TL', 'Kâr Marjı (%)': '% {:.2f}'
            }), use_container_width=True)
            
        with tab_siparis:
            st.write("API'den çekilen tüm siparişlerin satır satır Türkçe tarihli dökümü:")
            st.dataframe(tablayi_1den_baslat(df_sip).style.format({
                'Birim Fiyat': '{:,.2f} TL', 'Satış Tutarı (Ciro)': '{:,.2f} TL', 'Maliyet (TL)': '{:,.2f} TL',
                'Kargo (TL)': '{:,.2f} TL', 'Komisyon (%)': '% {:.2f}', 'Komisyon (TL)': '{:,.2f} TL',
                'Net Kâr (TL)': '{:,.2f} TL', 'Kâr Marjı (%)': '% {:.2f}'
            }), use_container_width=True)
            
        out_excel = BytesIO()
        with pd.ExcelWriter(out_excel, engine='openpyxl') as wr:
            tablayi_1den_baslat(ozet_t).reset_index().to_excel(wr, index=False, sheet_name='Ürün Bazlı Özet')
            tablayi_1den_baslat(df_sip).reset_index().to_excel(wr, index=False, sheet_name='Detaylı Sipariş Listesi')
            
            wb = wr.book
            for sh_name in wb.sheetnames:
                ws = wb[sh_name]; fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid"); font = Font(bold=True, color="FFFFFF")
                for col_idx, cell in enumerate(ws[1], 1): cell.fill = fill; cell.font = font
                for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width = 16
        out_excel.seek(0)
        
        st.download_button(
            label="📥 Satış ve Kârlılık Raporunu Excel Olarak İndir",
            data=out_excel,
            file_name=f"Trendyol_Satis_Analizi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# ==========================================
# SAYFA 6: AYARLAR & API BİLGİLERİ
# ==========================================
elif menu == "⚙ Ayarlar & API":
    st.markdown('<div class="main-title">⚙ Mağaza API ve Güvenlik Ayarları</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">API bilgilerinizi ve giriş şifrelerinizi buradan yönetebilirsiniz. Bilgiler sadece bilgisayarınızda (yerel) şifresiz olarak saklanır.</div>', unsafe_allow_html=True)
    
    api = load_api_settings()
    st.markdown("### 🟠 Trendyol API Anahtarları")
    with st.form("api_form"):
        ty_id = st.text_input("Satıcı ID (Seller ID)", value=api["ty_seller_id"])
        ty_key = st.text_input("API Key", value=api["ty_api_key"], type="password")
        ty_sec = st.text_input("API Secret", value=api["ty_api_secret"], type="password")
        btn_api = st.form_submit_button("💾 API Bilgilerini Kaydet", use_container_width=True)
        if btn_api:
            save_api_settings({"ty_seller_id": ty_id.strip(), "ty_api_key": ty_key.strip(), "ty_api_secret": ty_sec.strip()})
            st.success("✅ Trendyol API bilgileri başarıyla kaydedildi! Artık satış analizi sekmesinden tekrar bilgi girmenize gerek yoktur.")
            
    st.markdown("---")
    st.markdown("### 👥 Kullanıcı ve Şifre Yönetimi")
    
    tab_sifre, tab_yeni_user, tab_liste = st.tabs(["🔑 Şifre Değiştir", "➕ Yeni Kullanıcı Oluştur", "📋 Mevcut Kullanıcılar"])
    
    with tab_sifre:
        with st.form("sifre_degistir_form"):
            st.info(f"📍 Mevcut oturum açmış kullanıcı: **{st.session_state.get('current_user', 'Bilinmiyor')}**")
            eski_sifre = st.text_input("Mevcut Şifre", type="password")
            yeni_sifre = st.text_input("Yeni Şifre", type="password")
            yeni_sifre_tekrar = st.text_input("Yeni Şifre (Tekrar)", type="password")
            btn_sifre = st.form_submit_button("💾 Şifreyi Güncelle", use_container_width=True)
            if btn_sifre:
                curr_u = st.session_state.get('current_user', '')
                if auth_data["users"].get(curr_u) != eski_sifre.strip():
                    st.error("❌ Mevcut şifrenizi yanlış girdiniz!")
                elif len(yeni_sifre.strip()) < 3:
                    st.error("❌ Yeni şifre en az 3 karakter olmalıdır!")
                elif yeni_sifre.strip() != yeni_sifre_tekrar.strip():
                    st.error("❌ Yeni şifreler birbiriyle uyuşmuyor!")
                else:
                    auth_data["users"][curr_u] = yeni_sifre.strip()
                    save_auth(auth_data)
                    st.success("✅ Şifreniz başarıyla güncellendi! Sonraki girişinizde yeni şifreniz geçerli olacaktır.")

    with tab_yeni_user:
        with st.form("yeni_kullanici_form"):
            yeni_kadi = st.text_input("Yeni Kullanıcı Adı")
            yeni_kuser_sifre = st.text_input("Şifre", type="password")
            yeni_kuser_sifre_tekrar = st.text_input("Şifre (Tekrar)", type="password")
            btn_yeni_user = st.form_submit_button("➕ Kullanıcıyı Oluştur", use_container_width=True)
            if btn_yeni_user:
                k_adi_temiz = yeni_kadi.strip()
                if not k_adi_temiz:
                    st.error("❌ Kullanıcı adı boş bırakılamaz!")
                elif k_adi_temiz in auth_data["users"]:
                    st.error("❌ Bu kullanıcı adı zaten sistemde kayıtlı!")
                elif len(yeni_kuser_sifre.strip()) < 3:
                    st.error("❌ Şifre en az 3 karakter olmalıdır!")
                elif yeni_kuser_sifre.strip() != yeni_kuser_sifre_tekrar.strip():
                    st.error("❌ Şifreler birbiriyle uyuşmuyor!")
                else:
                    auth_data["users"][k_adi_temiz] = yeni_kuser_sifre.strip()
                    save_auth(auth_data)
                    st.success(f"✅ `{k_adi_temiz}` adlı yeni kullanıcı başarıyla oluşturuldu! Artık bu kullanıcı adı ve şifre ile de giriş yapılabilir.")

    with tab_liste:
        st.write("Sistemde erişim yetkisi olan kullanıcı adları aşağıda listelenmiştir:")
        user_list = list(auth_data["users"].keys())
        for idx, u in enumerate(user_list, 1):
            st.markdown(f"**{idx}.** 👤 `{u}` {'*(Aktif Oturum)*' if u == st.session_state.get('current_user') else ''}")
